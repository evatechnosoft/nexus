"""
Excel içe/dışa aktarma modülü.
- Export: Tüm envanter + talepler + atamalar → .xlsx
- Import: Excel'den cihaz listesi içe aktar
"""

import io
import json
from datetime import datetime

from fastapi import APIRouter, Depends, Request, UploadFile, File
from fastapi.responses import StreamingResponse, HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import get_db
from routers.requests import _get_or_create_person
import models

router = APIRouter(prefix="/excel", tags=["excel"])
templates = Jinja2Templates(directory="templates")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F9FAFB")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 4, 50
        )


@router.get("/export")
def export_excel(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Envanter ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Envanter"
    headers1 = [
        "ENV Kodu",
        "ID",
        "Tür",
        "Marka",
        "Model",
        "Seri No",
        "IMEI",
        "Durum",
        "Notlar",
        "Eklenme Tarihi",
    ]
    ws1.append(headers1)
    _style_header(ws1)

    devices = db.query(models.Device).order_by(models.Device.id).all()
    for i, d in enumerate(devices, start=2):
        inv_code = d.inventory_code or f"ENV-{d.id:04d}"
        row = [
            inv_code,
            d.id,
            d.device_type,
            d.brand,
            d.model,
            d.serial_no,
            d.imei or "",
            d.status,
            d.notes or "",
            d.created_at.strftime("%d.%m.%Y"),
        ]
        ws1.append(row)
        if i % 2 == 0:
            for cell in ws1[i]:
                cell.fill = ALT_FILL
    _auto_width(ws1)

    # ── Sheet 2: Talepler ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Talepler")
    headers2 = [
        "Talep ID",
        "Kişi Kodu",
        "Ad Soyad",
        "E-posta",
        "Telefon",
        "Adres",
        "Departman",
        "Cihaz Türü",
        "Durum",
        "Notlar",
        "Tarih",
    ]
    ws2.append(headers2)
    _style_header(ws2)

    requests_all = db.query(models.Request).order_by(models.Request.id).all()
    for i, r in enumerate(requests_all, start=2):
        person_code = (r.person.person_code if r.person else "") or ""
        row = [
            r.id,
            person_code,
            r.requester_name,
            r.requester_email,
            r.requester_phone or "",
            r.requester_address or "",
            r.department.name,
            r.device_type,
            r.status,
            r.notes or "",
            r.created_at.strftime("%d.%m.%Y"),
        ]
        ws2.append(row)
        if i % 2 == 0:
            for cell in ws2[i]:
                cell.fill = ALT_FILL
    _auto_width(ws2)

    # ── Sheet 3: Atamalar ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Atamalar")
    headers3 = [
        "ID",
        "Talep ID",
        "Alıcı",
        "Departman",
        "Cihaz",
        "Seri No",
        "Atama Tarihi",
        "Gönderim Tarihi",
        "Kargo No",
        "Notlar",
    ]
    ws3.append(headers3)
    _style_header(ws3)

    assignments = db.query(models.Assignment).order_by(models.Assignment.id).all()
    for i, a in enumerate(assignments, start=2):
        row = [
            a.id,
            a.request_id,
            a.request.requester_name,
            a.request.department.name,
            f"{a.device.brand} {a.device.model}",
            a.device.serial_no,
            a.assigned_at.strftime("%d.%m.%Y %H:%M"),
            a.shipped_at.strftime("%d.%m.%Y") if a.shipped_at else "",
            a.tracking_no or "",
            a.notes or "",
        ]
        ws3.append(row)
        if i % 2 == 0:
            for cell in ws3[i]:
                cell.fill = ALT_FILL
    _auto_width(ws3)

    # ── Sheet 4: Kişiler ──────────────────────────────────────────────────
    ws_p = wb.create_sheet("Kişiler")
    headers_p = [
        "Kişi Kodu",
        "Ad Soyad",
        "E-posta",
        "Telefon",
        "Adres",
        "Kayıt Tarihi",
        "Toplam Talep",
    ]
    ws_p.append(headers_p)
    _style_header(ws_p)

    persons = db.query(models.Person).order_by(models.Person.id).all()
    for i, p in enumerate(persons, start=2):
        row = [
            p.person_code or f"KSI-{p.id:04d}",
            p.name,
            p.email,
            p.phone or "",
            p.address or "",
            p.created_at.strftime("%d.%m.%Y"),
            len(p.requests),
        ]
        ws_p.append(row)
        if i % 2 == 0:
            for cell in ws_p[i]:
                cell.fill = ALT_FILL
    _auto_width(ws_p)

    # ── Sheet 5: Config Profilleri ─────────────────────────────────────────
    ws4 = wb.create_sheet("Config Profilleri")
    headers4 = ["Departman", "Kategori", "Config Adı", "Açıklama"]
    ws4.append(headers4)
    _style_header(ws4)

    depts = db.query(models.Department).all()
    for dept in depts:
        for i, item in enumerate(dept.config_items):
            row = [dept.name, item.category, item.name, item.description or ""]
            ws4.append(row)
    _auto_width(ws4)

    # ── Çıktı ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"it_envanter_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/import-page", response_class=HTMLResponse)
def import_page(request: Request):
    return templates.TemplateResponse("excel_import.html", {"request": request})


@router.post("/import")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Excel'den cihaz içe aktar.
    Beklenen sütunlar (Envanter sayfası):
    Tür | Marka | Model | Seri No | Notlar
    """
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))

    ws = wb.active  # İlk sayfa
    added = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        # Sütun sırası: Tür, Marka, Model, Seri No, IMEI (opsiyonel), Notlar
        try:
            device_type = str(row[0]).strip()
            brand = str(row[1]).strip()
            model = str(row[2]).strip()
            serial_no = str(row[3]).strip()
            imei = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            notes = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        except (IndexError, TypeError):
            skipped += 1
            continue

        if not serial_no:
            skipped += 1
            continue

        existing = db.query(models.Device).filter_by(serial_no=serial_no).first()
        if existing:
            skipped += 1
            continue

        if imei and db.query(models.Device).filter_by(imei=imei).first():
            skipped += 1
            continue

        device = models.Device(
            device_type=device_type,
            brand=brand,
            model=model,
            serial_no=serial_no,
            imei=imei or None,
            notes=notes or None,
            status=models.DeviceStatus.in_stock,
        )
        db.add(device)
        db.flush()
        device.inventory_code = f"ENV-{device.id:04d}"
        added += 1

    db.commit()
    return RedirectResponse(
        f"/inventory/?imported={added}&skipped={skipped}", status_code=303
    )


@router.post("/import-requests")
async def import_requests(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Excel'den personel taleplerini içe aktar.
    Beklenen sütunlar:
    Ad Soyad | E-posta | Telefon | Adres | Departman | Cihaz Türü | Notlar
    """
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    added = 0
    skipped = 0

    # Varsayılan departman (bulunamazsa kullanılacak)
    default_dept = db.query(models.Department).first()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:  # Ad ve E-posta zorunlu
            continue

        try:
            name = str(row[0]).strip()
            email = str(row[1]).strip()
            phone = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            address = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            dept_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            device_type = str(row[5]).strip() if len(row) > 5 and row[5] else "Laptop"
            notes = str(row[6]).strip() if len(row) > 6 and row[6] else ""

            # Departman bul
            dept = (
                db.query(models.Department)
                .filter(models.Department.name.ilike(dept_name))
                .first()
            )
            if not dept:
                dept = default_dept

            if not dept:
                skipped += 1
                continue

            person = _get_or_create_person(db, name, email, phone, address)

            new_req = models.Request(
                person_id=person.id,
                requester_name=person.name,
                requester_email=person.email,
                requester_phone=person.phone,
                requester_address=person.address,
                department_id=dept.id,
                device_type=device_type,
                notes=notes or None,
                status=models.RequestStatus.pending,
            )
            db.add(new_req)
            added += 1

        except Exception:
            skipped += 1
            continue

    db.commit()
    return RedirectResponse(
        f"/requests/?imported={added}&skipped={skipped}", status_code=303
    )
